import xmlrpc.client
import os
import pandas as pd
from openpyxl import Workbook, load_workbook

FILE_NAME = 'data.xlsx'
FIELDS = ['student_id', 'name', 'age']

def write_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(FIELDS)
        wb.save(FILE_NAME)

def add_student(student_id, name, age=None):
    if not os.path.exists(FILE_NAME):
        write_excel()
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([student_id, name, age])
    wb.save(FILE_NAME)
    print(f"Student {name} added successfully to Excel.")

def read_excel():
    if not os.path.exists(FILE_NAME):
        print(f"File {FILE_NAME} not found.")
        return pd.DataFrame(columns=FIELDS)
    df = pd.read_excel(FILE_NAME, engine='openpyxl')
    return df

# -----------------xml-rpc---------------------
url = "http://localhost:8888"

db = "test_db2"
username = "admin"
password = "admin"

common = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/common")
uid = common.authenticate(db, username, password, {})

if not uid:
    raise Exception("Authentication failed. Check your username/password/db name.")

models = xmlrpc.client.ServerProxy(f"{url}/xmlrpc/2/object")

def convert_excel():
    df = read_excel()
    for _, row in df.iterrows():
        student_id = row['student_id']
        name = row['name']
        age = row.get('age')
     
        existing = models.execute_kw(
            db, uid, password,
            'rest.student', 'search',
            [[['student_id', '=', student_id]]]
        )

        if existing:
            models.execute_kw(
                db, uid, password,
                'rest.student', 'write',
                [existing, {'name': name, 'age': age}]
            )
            print(f"Updated student {student_id} in Odoo.")
        else:
            models.execute_kw(
                    db, uid, password,
                    'rest.student', 'create',
                    [{'student_id': student_id, 'name': name, 'age': age}]
                )
            print(f"Created student {student_id} in Odoo.")

# -------------------- Main --------------------
if __name__ == '__main__':
    write_excel()
    add_student('101','Jay',32)
    # add_student('STD009', 'Om', 22)

    print(read_excel())

    convert_excel()





